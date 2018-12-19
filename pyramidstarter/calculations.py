#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Written for python 3, not tested under 2.
"""
"""
__author__ = "Matteo Ferla. [Github](https://github.com/matteoferla)"
__email__ = "matteo.ferla@gmail.com"
__date__ = ""

N = "\n"
T = "\t"
# N = "<br/>
import json
import os
import re
import uuid
import itertools
import operator

import numpy as np
import openpyxl
from Bio.Seq import Seq
from pyramid.response import FileResponse
from collections import Counter, defaultdict
from warnings import warn
from functools import reduce

from pyramidstarter import bike
from pyramidstarter.QQC import Trace, scheme_maker, codon_to_AA
from pyramidstarter.deep_mut_scanning import deep_mutation_scan
from pyramidstarter.mutagenesis import MutationTable, MutationDNASeq

from mako.template import Template

PATH = "/opt/app-root/src/pyramidstarter/"
if not os.path.isdir(PATH):
    PATH = "pyramidstarter/"
if __name__ == '__main__':
    PATH = ''


class SeqEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Seq):
            return str(obj)
        else:
            return json.JSONEncoder.default(self, obj)


def DS(req):
    ###seq.
    # for some reason I abided by the keep python variables pythonic and the JS variables JavaScribal
    req['sequence'] = re.sub('\>.*\n', '', req['sequence']).upper()  # kill fasta header
    req['sequence'] = re.sub('\W', '', req['sequence'])  # kill spaces and stuff
    #### check if user has given numeric or seq position
    if not re.search('\d', req['startBase']):
        req['startBase'] = re.sub('\W', '', req['startBase']).upper()
        if req['sequence'].find(req['startBase']) != -1:
            req['startBase'] = req['sequence'].find(req['startBase']) + 1
        else:
            pass
        if not re.search('\d', req['stopBase']):
            req['stopBase'] = re.sub('\W', '', req['stopBase']).upper()
            if req['sequence'].find(req['stopBase']) != -1:
                req['stopBase'] = req['sequence'].find(req['stopBase']) + 1 + len(req['stopBase'])
            else:
                pass
    #### parse rest
    pyreq = {
        'region': req['sequence'],
        'primer_range': [int(x) for x in req['primerRange'].split(',')],
        'section': slice(int(req['startBase']), int(req['stopBase'])),
        'Tm_bonus': float(req['TMBonus'])}
    if req['task'] == 'MP':
        pyreq['mutation'] = req['mutationList']
        pyreq['task'] = 'MP'
    elif req['task'] == 'DS':
        pyreq['mutation'] = req['mutationCodon']
        pyreq['task'] = 'DS'
    else:
        print(req)
        raise ValueError('Unrecognised: MP or DS mode?')
    if req['targetTemp']:
        pyreq['target_temp'] = float(req['targetTemp'])
    if req['GCclamp']:
        pyreq['GC_bonus'] = float(req['GCclamp'])
    cleanpyreq = {k: pyreq[k] for k in pyreq if pyreq[k]}
    cleanpyreq['staggered'] = req['method']
    data = deep_mutation_scan(**cleanpyreq)
    #####send to fxn
    fields = 'base AA codon fw_primer rv_primer len_homology fw_len_anneal rv_len_anneal fw_len_primer rv_len_primer homology_start homology_stop homology_Tm fw_anneal_Tm rv_anneal_Tm'.split()
    table = "<h3>Results</h3><table class='table-striped'>\n<thead><tr><th class='th-rotated'><div><span>" + "</span><div></th><th class='th-rotated'><div><span>".join(
        fields).replace('_', ' ') + "</span><div></th></tr></thead><tbody>\n"
    for row in data:
        for pref in ('fw', 'rv'):
            seq = row[pref + '_primer']
            n = 10
            row[pref + '_primer'] = ' '.join([str(seq[i:i + n]) for i in range(0, len(seq), n)])
        table += "<tr><td>" + "&nbsp;&nbsp;</td><td>".join([str(row[f]) for f in fields]) + "&nbsp;&nbsp;</td></tr>\n"
    table += "</tbody></table>"
    return json.dumps({'data': data, 'html': table}, cls=SeqEncoder)


def QQC(file_path, stored_filename, tainted_filename, location, scheme='NNK'):
    chroma = Trace.from_filename(file_path)
    Q = chroma.QQC(location, scheme)
    index = chroma.find_peak(location)
    span = round(len(chroma.A) / len(chroma.peak_index))
    window = (len(location) + 5) * span
    doubleindex = chroma.peak_index[index]
    raw = {
        base:
            [getattr(chroma, base)[doubleindex + i] for i in range(window)]
        for base in 'ATGC'}
    html = '''<div class="col-lg-12">&nbsp;<div id='QQC_raw_plot'></div></div><br/>
         <div class="col-lg-12">&nbsp;<p>Q<sub>pool</sub> = {q:.2f}</p></div><br/>
         <div class="col-lg-12">&nbsp;<div id='QQC_pie'></div></div><br/>
         <div class="col-lg-12">&nbsp;<div id='QQC_bar'></div></div><br/>
         '''.format(q=Q.Qpool)
    return json.dumps({'data': {'raw': raw, 'nt': Q.codon_peak_freq, 'AAemp': Q.empirical_AA_probabilities,
                                'AAscheme': Q.scheme_AA_probabilities, 'Qpool': Q.Qpool}, 'html': html})


def MC(file_path, stored_filename, tainted_filename, sequence, reverse=False, show=11, sigma=5):

    def prep_windows(variant):
        raws = []
        window_seq = []
        window_subseq = []
        diff = []
        pv_indices=[]
        for nti, mut in variant:
            # figure out why difference... backwards way as 2 mutations in the same codon are different than alone.
            diff.append([int(show / 2) + (1 + n - nti) for n in range(nti, nti + 3) if
                         neochroma.alignment[1][n] != neochroma.alignment[0][n]])
            # output the neighbourhood
            doubleindex = chroma.peak_index[nti]
            raws.append({base:
                             [getattr(neochroma, base)[doubleindex + i] for i in range(-window, window)]
                         for base in 'ATGC'})
            window_seq.append(
                ['{0}{1}{2}'.format(ref_AA[int(j / 3)], int(j / 3) + 1, query_AA[int(j / 3)]) if j % 3 == 1 else ' ' for
                 j in
                 range(nti - int(show / 2), nti + int(show / 2) + 1)])
            window_subseq.append(
                [str(1 + (j % 3)) for j in range(nti - int(show / 2), nti + int(show / 2) + 1)])
            pv_indices.append(str(int(nti/3)))
        return {'raw': raws,
                'mutants': [resn for (nti, resn) in variant],
                'codons': [(neochroma.alignment[1][nti:nti + 3], neochroma.alignment[0][nti:nti + 3]) for (nti, resn) in
                           variant],
                'differing': diff,
                'window_seq': window_seq,
                'window_subseq': window_subseq,
                'zeroed_indices_for_pv':pv_indices}

    chroma = Trace.from_filename(file_path)
    if isinstance(reverse, str):  # js sends a string...
        reverse = True if reverse == 'true' else False
    if reverse:
        neochroma = chroma.reverse().align(sequence)
    else:
        neochroma = chroma.align(sequence)
    ref_AA = Seq(neochroma.alignment[1].replace('-', '')).translate()
    query_AA = Seq(neochroma.alignment[0].replace('-', '')).translate()
    mutants = [(resi * 3, '{0}{1}{2}'.format(ref_AA[resi], resi + 1, query_AA[resi])) for resi in range(len(ref_AA)) if
               ref_AA[resi] != query_AA[resi]]
    noise = neochroma.noise_analysis(sigma=sigma)
    heteromutants = [(nti, '{0}{1}{2}'.format(ref_AA[int(nti/3)],int(nti/3)+1,query_AA[int(nti/3)])) for (nti, prop) in noise['outliers']]



    html = "<div class='row'><div class='col-lg-12'><p>There are {n} coding mutations ({l})</p></div></div>".format(n=len(mutants),
                                                                                l=' '.join([x[1] for x in mutants])) + \
           "<div class='row'>" + \
           "\n".join(["<div class='col-lg-6'>&nbsp;<div id='MC_mutant_{id}'></div></div>".format(id=i) for i in
                      range(len(mutants))]) + \
           "</div>" + \
           "<div id='MC_viewer'>&nbsp;</div>" +\
           "<div class='row'><div class='col-lg-12'>&nbsp;<div id='MC_noise'></div></div></div>" + \
           "\n".join(["<div class='col-lg-6'>&nbsp;<div id='MC_heteromutant_{id}'></div></div>".format(id=i) for i in
                      range(len(heteromutants))])
    # in QCC raw contains the snapshot. here it is a list of raw
    window = round(show / 2 * neochroma.span)
    return json.dumps({'data': {'window': window * 2,
                                'mutants': prep_windows(mutants),
                                'heteromutants': prep_windows(heteromutants),
                                'noise': noise},
                       'html': html})


def glue(jsonreq):
    # {'prob_complete': '0.95', 'completeness': '0.95', 'library_size': '3000000', 'nvariants': '1000000', 'mode': 'prob_complete'}
    for id in ['prob_complete', 'completeness', 'library_size']:
        if jsonreq['mode'] == id:
            glue_out = bike.glue(jsonreq['nvariants'], **{id: jsonreq[id]})
            if float(jsonreq['nvariants']) < 10:
                hreply = '<div class="alert alert-info" role="alert"><i class ="fa fa-exclamation-triangle" aria-hidden="true" > </ i > <span>Number of variants very small. Poisson statistics may be compromised.</span></div>'
            else:
                hreply = '<div></div>'
            return json.dumps({'data': glue_out, 'html': hreply})


def glueit_csh(jsonreq):
    """
    This version runs the cshell which does not work in Linux.
    :param jsonreq:
    :return:
    """
    library_size=int(jsonreq['library_size']) #it will be stringified anyway. Not sure why I am converting it.
    filename = os.path.join(PATH, 'tmp', '{0}.txt'.format(uuid.uuid4()))
    with open(filename, 'w') as f:
        for i in range(1,int(jsonreq['num_codons'])):
            f.write(jsonreq['codon{}'.format(i)]+'\n')
    #print('filename',filename)
    html=bike.glueit(library_size=library_size, codonfile=filename)
    return json.dumps({'data': None, 'html': html})

def glueit(jsonreq):
    filename = os.path.join(PATH, 'tmp', '{0}.dat'.format(uuid.uuid4()))
    with open(filename, 'w') as f:
        library_size=int(jsonreq['library_size'])
        num_codon=int(jsonreq['num_codons'])
        f.write('{nc} {ls}\n'.format(nc=num_codon,ls=library_size))
        for ci in range(1,6):
            if 'codon'+str(ci) in jsonreq and jsonreq['codon'+str(ci)] and jsonreq['codon'+str(ci)].find('-'):
                codon=jsonreq['codon'+str(ci)]
                # "NWH": {"V": 3, "*": 1, "H": 2, "I": 3, "E": 1, "N": 2, "Y": 2, "Q": 1, "K": 1, "L": 4, "D": 2, "F": 2}
                counts=Counter([codonball[codon][a] for a in codonball[codon] if a != '*'])
                f.write(' '.join([str(counts[i]) for i in range(1,7)])+'\n')
    html=bike.glueit(library_size=library_size, datfile=filename)
    return json.dumps({'data': None, 'html': html})

def pedelAA(jsonreq):
    # Super hacky for now. I struggled with the maths in python mode.
    filename = os.path.join(PATH, 'tmp', '{0}'.format(uuid.uuid4()))
    #sequence file
    seq=re.sub('[^ATGC]','',jsonreq['sequence'].upper().replace('U','T'))
    assert not len(seq) % 3, 'Sequence is not a multiple of three.'
    assert len(seq) < 50000, 'Sequence is longer than 50kb...'
    assert len(seq) > 0, 'Sequence cannot be empty'
    #TODO assert internal stop codons...
    # fix naming issue!
    jsonreq['nsubst']=jsonreq['load']
    jsonreq['ninsert'] = 0
    jsonreq['distr'] = 1 # 1 = Poisson, 0 = PCR
    jsonreq['ndelete'] = 0
    jsonreq['library_size']=jsonreq['size']
    jsonreq['nucnorm'] = int(jsonreq['nucnorm'])
    #TODO get values from JS!!!
    jsonreq['ncycles'] = 30
    jsonreq['eff'] = 0.8
    with open(filename + '.fasta', 'w') as f:
        f.write('>inseq\n{}\n'.format(seq))
    #setup file
    with open(filename+'.setup','w') as f:
        f.write(' \n'.join([filename+'.fasta',
                      filename + '.nuc.dat',
                      os.path.join(PATH, 'bikeshed','aa2codon.dat'),
                      os.path.join(PATH, 'bikeshed', 'Acodon.dat'),
                      filename+'.html',
                      filename + 'matrix.html',
                      filename + 'table.html',
                      filename + '.seqstats.txt']))
        f.write(' \n')
        f.write(' \n'.join([str(jsonreq[k]) for k in ['nsubst','ninsert','ndelete','library_size','nucnorm','distr','ncycles','eff']]))
        # Make nuc matrix: note pedel-AAc.cxx sets the diagonals = 0 anyway.
        pbases='T','C','A','G'  # this differs from elsewhere here.
        with open(filename + '.nuc.dat', 'w') as f:
            f.write('\n'.join([' '.join([str(jsonreq[origin+'2'+destination]) if origin!=destination else '0' for destination in pbases]) for origin in pbases ]))
    #test file...
    #filename = 'pyramidstarter/tmp/56a80370-c6d3-4827-adbe-5cfbeecf39dc'
    #warn('To test, the wrapper is circumvented.')
    data=bike.pedelAA(filename + '.setup')
    #print(data['html'])
    html=str(open(os.path.join(PATH,'templates','pedelAA_results.mako')).read()).format(**data)
    return json.dumps({'data': data['sub_table_data'], 'html': html})


def pedel(jsonreq):
    pedel_out = bike.pedel(library_size=jsonreq['size'], sequence_length=jsonreq['len'],
                           mean_number_of_mutations_per_sequence=jsonreq['mutload'])
    stats = bike.pedel_stats(library_size=jsonreq['size'], sequence_length=jsonreq['len'],
                             mean_number_of_mutations_per_sequence=jsonreq[
                                 'mutload'])
    # print(pedel_out)
    return json.dumps({'data': stats, 'html': '<div>Expected number of distinct sequences in library: {0}</div>'.format(
        pedel_out['expected_number_of_distinct_sequences_in_library'])})

def driver(jsonreq):
    """
    nasty...
    from {'length': '1425', 'library_size': '1600', 'mean': '2', 'positions' : '250 274 375 650 655 757 763 982 991'}
    library_size sequence_length mean_number_of_crossovers_per_sequence list_of_variable_positions_file outfile xtrue
    :param jsonreq:
    :return:
    """
    mode=jsonreq['mode']
    if isinstance(mode, str):  # js sends a string...
        mode = True if mode == 'true' else False
    if mode:
        a=re.sub('\W','',jsonreq['sequenceA'])
        b=re.sub('\W','',jsonreq['sequenceB'])
        print(a)
        print(b)
        if len(a) != len(b):
            raise ValueError('The lengths of the sequences differ. In future I might do an alignment, for now it&#8217;s matching')
        pos=[str(i+1) for i in range(len(a)) if a[i] != b[i]]
    else:
        pos = jsonreq['positions'].split()
    xtrue=jsonreq['xtrue']
    if isinstance(xtrue, str):  # js sends a string...
        xtrue = '1' if xtrue == 'true' else '0'
    elif isinstance(xtrue, bool):  # unneccassry here
        xtrue = '1' if xtrue == True else '0'
    else:
        pass # most likely a number already.
    listfile =os.path.join(PATH, 'tmp', 'listfile_{0}.txt'.format(uuid.uuid4()))
    outfile = os.path.join(PATH, 'tmp', 'outfile_{0}.txt'.format(uuid.uuid4()))

    open(listfile,'w').write('\n'.join([str(len(pos)),*pos]))
    driver_out=bike.driver(sequence_length=jsonreq['length'],
                           library_size=jsonreq['library_size'],
                           mean_number_of_crossovers_per_sequence=jsonreq['mean'],
                           list_of_variable_positions_file=listfile,
                           outfile=outfile,
                           xtrue=xtrue)
    hreply ="<p>"+driver_out + "</p>"+open(outfile).read()
    return json.dumps({'data': {'positions': ' '.join(pos)}, 'html':hreply})


def platecounter(size):
    if size == 96:
        letters = 'ABCDEFGH'
        m = 12
    elif size == 384:
        letters = 'ABCDEFGHIJKLMNOP'
        m = 24
    else:
        raise Exception('only 96 or 384 supported')
    for i in range(1, m + 1):
        for l in letters:
            yield l + str(i)

def silico(jsonreq):
    #get data
    load=float(jsonreq['load'])
    seq=MutationDNASeq(jsonreq['sequence'].upper())
    spectra=MutationTable({k: float(v) for (k,v) in jsonreq.items() if k.find('2') != -1}).normalize()
    ## randomise!
    bases=('A','T','G','C')
    n_muts=np.random.poisson(load) # n of mutations
    m_data=[]
    try:
        for mi in range(n_muts):
            # choose mutation type
            q=np.random.uniform(0,1)
            for m_type in spectra:
                q = q - spectra[m_type]
                if q < 0:
                    maps = {b: [i for i in range(len(seq)) if seq[i] == b] for b in bases} #has to be recalculated due to mut on mut
                    nti=maps[m_type[0]][np.random.random_integers(0,len(maps[m_type[0]]))]
                    m_data.append({'type':m_type,'nti':nti})
                    seq.mutate(str(nti+1)+m_type)
                    break
    except IndexError: #this happens due to the violation of the assumption of equal base prob... FIX!
        return silico(jsonreq)
    hreply ='<pre><code>>variant_'+' '.join([str(m) for m in sorted(seq.mutations, key=lambda x: x.num_aa)])+'\n'+str(seq)+'</pre></code>'
    return json.dumps({'data': None, 'html':hreply})

(codonball, codontallyball)=json.load(open(os.path.join(PATH,'AAcalc.json'),'r'))
AAnamemask = {'R': 'R', 'Ser': 'S', 'W': 'W', 'His': 'H', 'Isoleucine': 'I', 'Tryptophan': 'W', 'Thr': 'T', 'I': 'I',
              'Cys': 'C', 'Gln': 'Q', 'Y': 'Y', 'Methionine': 'M', 'Proline': 'P', 'K': 'K', 'S': 'S', 'Valine': 'V',
              'Glutamine': 'Q', 'Tyr': 'Y', 'Val': 'V', 'Histidine': 'H', 'Lysine': 'K', 'Glutamate': 'E', 'L': 'L',
              'Alanine': 'A', 'N': 'N', 'Gly': 'G', 'D': 'D', 'Ala': 'A', 'Glu': 'E', 'Phenylalanine': 'F', 'Met': 'M',
              'V': 'V', 'Glycine': 'G', 'Arg': 'R', 'A': 'A', 'M': 'M', 'Q': 'Q', 'Lys': 'K', 'Asparagine': 'N',
              'Pro': 'P', 'Trp': 'W', 'E': 'E', 'H': 'H', 'Tyrosine': 'Y', 'P': 'P', 'Cysteine': 'C', 'aspartate': 'D',
              'G': 'G', 'F': 'F', 'Asn': 'N', 'Leucine': 'L', 'Asp': 'D', 'Threonine': 'T', 'C': 'C', 'Phe': 'F',
              'T': 'T', 'Ile': 'I', 'Arginine': 'R', 'Serine': 'S', 'Leu': 'L'}
def codonAA(jsonreq):
    # parse
    def codon_input_parser(text):
        if not text:
            return {'!'}  #not sure why there is a bug where if no anti-codon is given it crashes, but that is the case.
        try:
            cleanset = {AAnamemask[aa] for aa in text.replace(',', ' ').replace(';', ' ').split()}
        except KeyError:  #the user did not space it...
            cleanset = {AAnamemask[aa] for a in text.replace(',', ' ').replace(';', ' ') for aa in a}
        return cleanset
    AAset=codon_input_parser(jsonreq['list'])
    antiAAset = codon_input_parser(jsonreq['antilist'])
    if AAset.intersection(antiAAset):
        return json.dumps({'data': None, 'html': '<div class="alert alert-danger" role="alert"><b>The Schr&ouml;dinger residue:</b> You cannot request both to have and not to have a residue at the same time...</div>'})
    validcodon=[]
    for cd in codonball:
        if AAset.issubset(set(codonball[cd].keys())) and not antiAAset.issubset(set(codonball[cd].keys())):
            #detemrining number of correct hits
            validcodon.append({'codon': cd,
                        'N_codons': sum(codonball[cd].values()),
                        'N_AA': len(set(codonball[cd].keys())),
                        '%_valid': round(100*sum([codonball[cd][aa] for aa in codonball[cd].keys() if aa in AAset])/sum(codonball[cd].values())) ,
                        '%_stop': round(100*codonball[cd]['*']/sum(codonball[cd].values()) if '*' in codonball[cd] else 0),
                        'N_selected': {aa: codonball[cd][aa] for aa in AAset},
                        'N_all': codonball[cd]
                        })
    # sort with a 2x penalty for stop and give precence to more equal distributions, in case of a tie to less degenerate codons.
    def varmod(x):
        mu=sum(x.values())/len(x)
        return 0.001* sum([(x[i] - mu) ** 2 for i in x])/len(x)
    svalidcodon=sorted(validcodon, key=lambda x: x['%_valid'] - 2* x['%_stop'] - varmod(x['N_selected']) - 0.0001*x['N_codons'], reverse=True)
    headers=['codon','N_codons','N_AA','%_valid','%_stop','N_selected','N_all']
    hreply='<table class=\'table table-striped\'><tr><th>{header}</th></tr><tr>{body}</tr></table>'.format(
        header='</th><th>'.join(headers),
        body='</tr><tr>'.join([''.join(['<td>{}</td>'.format(str(cdball[key])) for key in headers]) for cdball in svalidcodon[0:50]])
    )
    return json.dumps({'data': None, 'html': hreply})


def IDT(request, size):
    data = json.load(open(request.session['DS']))['data']
    wb = openpyxl.Workbook()
    platenumb = 1
    wf = wb.active
    wf.title = 'forward_{}'.format(platenumb)
    wr = wb.create_sheet('reverse_{}'.format(platenumb))
    wf['A1'] = 'Well Position'
    wf['B1'] = 'Name'
    wf['C1'] = 'Sequence'
    wr['A1'] = 'Well Position'
    wr['B1'] = 'Name'
    wr['C1'] = 'Sequence'
    '''{'fw_primer': 'AGGGCTCGAG CnnkGTGAGC AAGGGCG', 'fw_anneal_Tm': 55.7, 'fw_len_primer': 27,
     'rv_primer': 'TTGCTCACmn nGCTCGAGCC CTGG', 'rv_anneal_Tm': 57.4, 'homology_stop': 116, 'len_homology': 22,
     'rv_len_anneal': 13, 'rv_len_primer': 24, 'base': 105, 'AA': 'M1', 'homology_Tm': 60.5, 'homology_start': 94,
     'fw_len_anneal': 16, 'codon': 'ATG'}'''
    platewells = platecounter(size)
    r = 1
    for entry in data:
        r = r + 1
        try:
            well = next(platewells)
        except StopIteration:
            platenumb += 1
            wf = wb.create_sheet('forward_{}'.format(platenumb))
            wr = wb.create_sheet('reverse_{}'.format(platenumb))
            wf['A1'] = 'Well Position'
            wf['B1'] = 'Name'
            wf['C1'] = 'Sequence'
            wr['A1'] = 'Well Position'
            wr['B1'] = 'Name'
            wr['C1'] = 'Sequence'
            r = 2
            platewells = platecounter(size)
            well = next(platewells)
        wf.cell(row=r, column=1, value=well)
        wf.cell(row=r, column=2, value=entry['AA'])
        wf.cell(row=r, column=3, value=entry['fw_primer'].replace(' ', ''))
        wr.cell(row=r, column=1, value=well)
        wr.cell(row=r, column=2, value=entry['AA'])
        wr.cell(row=r, column=3, value=entry['rv_primer'].replace(' ', ''))
    request.session['DS_IDT'] = request.session['DS'].replace('.json', '.xlsx')
    wb.save(request.session['DS_IDT'])
    response = FileResponse(
        request.session['DS_IDT'],
        request=request,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    return response


def codon(request):
    ##parts are stolen from the QCC class. This basically just give the empirical results.
    # convert str request which should have codons to AA.
    scheme = scheme_maker(request)
    # a LIST of n primers each being a list of 3 positions each with a dictionary of the probability of each base.
    schemeprobball = []
    for primer in scheme:
        schemeprobball.append(codon_to_AA(primer[1]))
    AA = {
        aa: sum([scheme[pi][0] * schemeprobball[pi][aa] for pi in range(len(schemeprobball))]) for aa in
        schemeprobball[0].keys()}
    return json.dumps({'data': AA, 'html': ''})


def make_AAcalc_json():
    degeneracy = {'N': 'ATGC',
                  'A': 'A',
                  'T': 'T',
                  'G': 'G',
                  'C': 'C',
                  'W': 'AT',
                  'S': 'GC',
                  'K': 'GT',
                  'M': 'AC',
                  'R': 'AG',
                  'Y': 'TC',
                  'B': 'TGC',
                  'V': 'AGC',
                  'H': 'ATC',
                  'D': 'ATG'}
    codonball = dict()
    codontallyball = dict()
    for cd in itertools.product('ATGCWSKMYRBVHDN', repeat=3):
        incodonball = defaultdict(int)
        t = 0
        for cdx in itertools.product(degeneracy[cd[0]], degeneracy[cd[1]], degeneracy[cd[2]]):
            aa = str(Seq(''.join(cdx)).translate())
            incodonball[aa] = incodonball[aa] + 1
            t = t + 1
        codonball[''.join(cd)] = dict(incodonball)
        codontallyball[''.join(cd)] = t
    json.dump((codonball, codontallyball), open('AAcalc.json', 'w'))

def probably(jsonreq):
    load = float(jsonreq['load'])
    mlist=jsonreq['list']
    lib=float(jsonreq['size'])
    seq = MutationDNASeq(jsonreq['sequence'].upper())
    spectra = MutationTable({k: float(v) for (k, v) in jsonreq.items() if k.find('2') != -1}).normalize()
    mutant=seq.mutate(mlist)
    diff=(''.join([m.from_nuc for m in mutant.mutations]) ,''.join([m.to_nuc for m in mutant.mutations]))
    muts=['{f}>{t}'.format(f=m.from_nuc[n],t=m.to_nuc[n]) for m in mutant.mutations for n in range(len(m.to_nuc))]
    singles=[spectra[m] for m in muts]
    p=reduce(operator.mul, singles, 1)*load/len(seq)
    plib=1-(1-p)**lib
    return json.dumps({'data': {'p': p, 'mutations':muts}, 'html': '<h3>Results</h3><p>overall probability of encountering a {v} genotype in a given variant is {p:2g}, amid the library {l:2g}.</p>'.format(p=p, v=mlist, l=plib)})

def pmut_renumber(scores, AAlphabet='A C D E F G H I K L M N P Q R S T V W Y',delete=[], insert=None):
    if isinstance(AAlphabet, str):
        AAlphabet=AAlphabet.split()
    first_delete=[]
    last_delete=[]
    if delete:
        for d in delete:
            first_delete.append(int(d.split(':')[0]))
            last_delete.append(int(d.split(':')[1]))
    if insert:
        (first_insert, insert_amount)=[int(x) for x in insert.split('+')]
    else:
        (first_insert, insert_amount)= None, None

    if isinstance(scores, str):
        scores=scores.split('\n')
    #headers='resi from_resn to_resn average_ddG average_total_energy'.split()

    mutdex=defaultdict(dict)
    seq=dict()
    for line in scores:
        if isinstance(line, bytes):
            line=line.decode("utf-8")
        parts=line.split()
        #print('THIS IS A LINE',parts)
        if len(parts) < 2 or len(parts[1]) < 2 or parts[1][1] != '-': #the hyphen after chain...
            continue
        resi=int(parts[2][3:-1]) # resi is from PDB numbering.
        if resi and resi > 0 and resi < 1e3:
            #mutdex[resi].append([parts[2][2],parts[2][-1], parts[3], parts[4]]) # resi from_resn to_resn average_ddG average_total_energy
            mutdex[resi][parts[2][-1]]=float(parts[3])
            seq[resi]=parts[2][2]
            if parts[2][2] not in mutdex[resi]:
                mutdex[resi][parts[2][2]] = 0 # no mutation case.
    #fix missing
    mutdex={resi: {resn: 999 if resn not in mutdex[resi] else mutdex[resi][resn]  for resn in AAlphabet} for resi in range(1,max(mutdex.keys())+1)}
    seq={resi: 'X' if resi not in seq else seq[resi] for resi in range(1,max(mutdex.keys())+1)}
    #print(mutdex)
    #fix numbering.
    delete_flag = False
    current_offset = 0
    outdex={}
    outseq={}
    for resi in range(1,9999):
        if len(mutdex) < resi + current_offset:
            break
        elif delete_flag == True or resi in first_delete:
            if resi in last_delete: # last one to skip...
                delete_flag = False
            else:
                delete_flag = True
            current_offset -= 1
            #print('Deleting...', resi, mutdex[resi][0][0])
            continue
        elif resi == first_insert:
            outdex[resi]=mutdex[resi+current_offset]
            outseq[resi]=seq[resi+current_offset]
            for n in range(insert_amount):
                current_offset += 1
                outdex[resi] ={resn: 0 for resn in AAlphabet}
        elif not resi in mutdex: # no offsetting... just filling
            outdex[resi] = {resn: 0 for resn in AAlphabet}
            outseq[resi] = seq[resi + current_offset]
        else:
            outdex[resi] = mutdex[resi + current_offset]
            outseq[resi] = seq[resi + current_offset]
    return outdex, outseq

def delete_temp():
    for file in os.listdir(os.path.join(PATH,'tmp/')):
        os.remove(os.path.join(PATH,'tmp',file))

if __name__ == "__main__":
    #driver({'positions': '250 274 375 650 655 757 763 982 991', 'mean': '2', 'xtrue': True, 'library_size': '1600', 'length': '1425'})
    #pprint(codonAA({'list':'G P S'}))
    #bases='A','T','G','C'
    #print(probably({'sequence':'ATGGGCCCGAAATAG','size':1e6,'list':['M1A'],'load':5, **{b1+'>'+b2:8.333333333 for b1 in bases for b2 in bases}}))
    #print(glueit({'num_codons': 6, 'library_size':1000,'codon1':'ATG','codon2':'NNT','codon3':'NNK','codon4':'NNK','codon5':'NNK','codon6':'ATG'}))
    #print(pedelAA({'size':10000,'ninsert': 0, 'ndelete': 0,'load': 5,'A2T': '5', 'A2G': '8', 'A2C': '5', 'T2A': '14', 'T2G': '0', 'T2C': '5', 'G2A': '9', 'G2T': '4', 'G2C': '2', 'C2A': '3', 'C2T': '6', 'C2G': '3','nucnorm':0,'distr':'Poisson','ncycles':30,'eff':0.8,'sequence':'ATGGTGAGCAAGGGCGAGGAGCTGTTCACCGGGGTGGTGCCCATCCTGGTCGAGCTGGACGGCGACGTAAACGGCCACAAGTTCAGCGTCCGCGGCGAGGGCGAGGGCGATGCCACCAACGGCAAGCTGACCCTGAAGTTCATCTGCACCACCGGCAAGCTGCCCGTGCCCTGGCCCACCCTCGTGACCACCTTCGGCTACGGCGTGGCCTGCTTCAGCCGCTACCCCGACCACATGAAGCAGCACGACTTCTTCAAGTCCGCCATGCCCGAAGGCTACGTCCAGGAGCGCACCATCTCTTTCAAGGACGACGGTACCTACAAGACCCGCGCCGAGGTGAAGTTCGAGGGCGACACCCTGGTGAACCGCATCGAGCTGAAGGGCATCGACTTCAAGGAGGACGGCAACATCCTGGGGCACAAGCTGGAGTACAACTTCAACAGCCACTACGTCTATATCACGGCCGACAAGCAGAAGAACTGCATCAAGGCTAACTTCAAGATCCGCCACAACGTTGAGGACGGCAGCGTGCAGCTCGCCGACCACTACCAGCAGAACACCCCCATCGGCGACGGCCCCGTGCTGCTGCCCGACAACCACTACCTGAGCCATCAGTCCAAGCTGAGCAAAGACCCCAACGAGAAGCGCGATCACATGGTCCTGCTGGAGTTCGTGACCGCCGCCGGGATTACACATGGCATGGACGAGCTGTACAAGTAA'}))
    #print(codonAA({'list': 'ASR', 'antilist': ''})[70:])
    #print(codonAA({'list': 'ASR', 'antilist': 'C'})[70:])
    #print(pmut_renumber(open('pyramidstarter/static/dog-intermediate_scores.txt','r').read()))
    #delete_temp()
    pass